from typing import Annotated
from fastapi import APIRouter, Depends
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from typing import List  
# from .database import SessionLocal, engine

from sql_app.crud import get_device_info, get_location_info, get_financial_services_apps
from sql_app.database import SessionLocal
from sql_app.models import DeviceInfo, Location, FinancialServiceApp

router = APIRouter()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

db_dependency = Annotated[Session, Depends(get_db)]


@router.get("/report")
async def pdf_generator(db: db_dependency):
    workbook = Workbook()

    #* Finalizado
    def build_device_info_sheet(db: Session): 
        sheet = workbook.create_sheet("Devices")
        sheet.append(["clientDeviceID", "OSName", "internalmodelname", "osversion", "lastUpdateDate", "insertionDate", "cpumodel", "TotalMemory", "ExternalStorage"])
        items_from_DB : List[DeviceInfo] = get_device_info(db)
        for i in items_from_DB:
            item_row = [i.device_id , i.operative_system, i.model , i.so_version, "", "", i.cpu ]
            sheet.append(item_row) 
    # TODO: requiere info de API google maps o mejorar modelo que viene de la app
    def build_poi_info_sheet(db: Session): 
        sheet = workbook.create_sheet("POI_per_device")
        items_from_DB : List[Location] = get_location_info(db)
        sheet.append(["deviceId", "brand_name", "name_default", "category_all", "county", "region", "latitude", "longitude"])
        for i in items_from_DB:
            item_row = [ "identifier" , "brand_name" , i.administrative_area, "category" , i.country, i.locality, i.latitute , i.longitude ]
            sheet.append(item_row) 

    def build_finance_services_sheet(db:Session):
        sheet = workbook.create_sheet("Finance Services per Devices")
        items_from_DB: List[FinancialServiceApp]= get_financial_services_apps(db)
        sheet.append(["device_id", "os_name", "package_name"])
        for i in items_from_DB:
            item_row = [i.device_id, i.os_name, i.package_name]
            sheet.append(item_row)
    
    #!realizar lo del bluetooth con mac del dispositivo
    def build_connected_devices_sheet(db: Session):
        sheet = workbook.create_sheet("Connected Device Preferences")
        sheet.append(["deviceId", "Functionality", "Brand"])
        pass
            


    # sheet_devices = wb.create_sheet("Devices")
    # sheet_devices.append(["clientDeviceID", "OSName", "internalmodelname", "osversion", "lastUpdateDate", "insertionDate", "cpumodel", "TotalMemory", "ExternalStorage"])

    # sheet_poi = wb.create_sheet("POI_per_device")
    # sheet_poi.append(["deviceId", "brand_name", "name_default", "category_all", "county", "region", "latitude", "longitude"])

    # sheet_bt_devices = wb.create_sheet("Connected Device Preferences")
    # sheet_bt_devices.append(["deviceId", "Functionality", "Brand"])

    # sheet_bt_devices = wb.create_sheet("Finance Services Apps")
    # sheet_bt_devices.append(["Online Financial Service Uid", "Count of device_id"])

    # sheet_bt_devices = wb.create_sheet("Network Services Used")
    # sheet_bt_devices.append(["deviceId", "type", "percent", "day_range"])


    build_device_info_sheet(db)
    build_poi_info_sheet(db)
    build_finance_services_sheet(db)
    build_connected_devices_sheet(db)

    filename = "report.xlsx"
    workbook.save(filename)

    # return FileResponse(filename) #lo abre en el navegador
    return FileResponse(filename, filename = filename) #lo descarga como archivo